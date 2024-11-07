import os
import streamlit as st
import pandas as pd
from datetime import datetime

def get_file_list(suffix, path):
    """
    获取当前目录所有指定后缀的文件名列表和绝对路径列表
    :param suffix: 文件后缀
    :param path: 搜索路径
    :return: 文件名列表, 文件路径列表
    """
    file_names = []
    file_paths = []
    for root, _, files in os.walk(path, topdown=False):
        for name in files:
            if os.path.splitext(name)[1] == suffix:
                file_names.append(name)
                file_paths.append(os.path.join(root, name))
    return file_names, file_paths

def page_three():
    st.title("干线装载率早晚报")

    file = st.sidebar.file_uploader("上传文件", type=['csv', 'xlsx'])

    # 添加时间选择器
    selected_time = st.sidebar.time_input("选择筛选时间", value=datetime.strptime("09:00", "%H:%M").time())

    if file is not None:
        # 读取 CSV 文件
        df = pd.read_excel(file)

        # 将重量装载率从百分比转换为小数
        def convert_percentage_to_decimal(percentage_str):
            try:
                return float(percentage_str.strip('%')) / 100
            except ValueError:
                return None
        
        df['重量装载率'] = df['重量装载率'].apply(convert_percentage_to_decimal)


        # 确保线路里程列能被转换成数值类型
        df['线路里程（KM）'] = pd.to_numeric(df['线路里程（KM）'], errors='coerce')



        # 数据清洗：删除中转重量=0的数据
        df = df[df['计费重量(kg)'] != 0]

        excluded_demand_types = ['物料配送-调拨', '编织袋调拨', '笼车调拨', '更改卸货场地']
        df = df[~df['需求类别'].isin(excluded_demand_types)]

        excluded_vehicle_types = ['粤港常温厢式运输车', '粤澳常温厢式运输车']
        df = df[~df['计划车辆类型'].isin(excluded_vehicle_types)]

        '''excluded_area_types = ['020RH','020WS','020WK','755RH','755VF','769WK','769WS','757WH','757WL','660VH','663WH','752WH','762VH','668VH','760WH','771WH','772VH','8981VH','898WH']
        df = df[~df['发出网点'].isin(excluded_area_types)]'''

        included_area_types = ['769XG','769WM','769WF','769WB','758CC001','757WA','757WE','763X','020WM','020RD','020W','020WE','020XB','755X','755WM','755WJ','755WE','755W','755WF','755RC','663WA','663WB','752WA','762VA','750VA','750W','760W','756W','759VA']
        df = df[df['发出网点'].isin(included_area_types)]

        # 将计划发车时间转换为日期时间类型
        df['计划发车时间'] = pd.to_datetime(df['计划发车时间'])
        df['实际发车时间'] = pd.to_datetime(df['实际发车时间'])


        # 假设运输等级列名为 "运输等级"
        df['运输等级'] = df['运输等级'].fillna('未知')  # 处理缺失值

        # 将计划发车日期转换为日期时间类型
        df['日期'] = pd.to_datetime(df['日期'])

        # 定义筛选条件：计划发车时间早于对应日期的10:30
        def is_earlier_than_1030(row):
            target_time = datetime(row['日期'].year, row['日期'].month, row['日期'].day,selected_time.hour, selected_time.minute)
            return row['实际发车时间'] <= target_time

        df = df[df.apply(is_earlier_than_1030, axis=1)]

        # 定义片区划分规则
        def assign_region(code):
            region_map = {
                '东莞片': ['769WB', '769WF', '769WM', '769XG', '769WV', '769VZ',],
                '佛山片': ['757WA', '757WL', '757WE', '758CC001', '763X', '763XD'],
                '广州片': ['020RD', '020W', '020WG', '020WM',  '020XB', '020CC006', '020CC001', '020R', '020CC008', '020XK', '020WT', '020Z', '020WE', '020VM'],
                '深圳片': [ '755W', '755WE', '755WF', '755WM', '755X', '755R', '755RC', '755VM', '755Z', '755WJ', '755ZS'],
                '粤东片': ['663WA', '752WA',  '762VA', '663R', '768WA', '663WB', '768VA',  '660VW'],
                '粤西片': ['750W', '756W', '760W', '750VA',  '756R', '759VA', '750WA', '668VA']
            }
            for region, codes in region_map.items():
                if code in codes:
                    return region
            return '未知'

        # 应用片区划分规则
        df['片区'] = df['发出网点'].apply(assign_region)

        # 定义固定的片区顺序列表
        fixed_regions = ['东莞片', '佛山片', '广州片', '深圳片', '粤东片', '粤西片']

        # 定义运输等级
        transport_levels = ['一级运输', '二级运输']

        # 获取所有日期
        all_dates = df['日期'].dt.date.unique()
        all_dates = sorted(all_dates)

        # 按照片区、日期和运输等级分类计算发运货量、发运票数、发运车次、低装载车次、装载率
        results_data1 = []

        for date in all_dates:
            date_df = df[df['日期'].dt.date == date]
            
            
            # 总体数据
            total_weight = date_df['计费重量(kg)'].sum()
            total_tickets = date_df['应到(件)'].sum()
            total_trips = date_df['计划需求ID'].nunique()
            total_full_weight = date_df['满载重量(kg)'].sum()
            low_load_trips = date_df[((date_df['线路里程（KM）'] > 800) & (date_df['重量装载率'] < 0.7)) | ((date_df['线路里程（KM）'] <= 800) & (date_df['重量装载率'] < 0.5))]['计划需求ID'].nunique()
            loading_rate = total_weight / total_full_weight if total_full_weight > 0 else 0

            # 存储总体数据
            results_data1.append({
                '日期': date,
                '片区': '分拨区',
                '运输等级': '总计',
                '发出网点': '总计',
                '发运货量': total_weight,
                '满载重量': total_full_weight,
                '发运票数': total_tickets,
                '发运车次': total_trips,
                '低装载车次': low_load_trips,
                '装载率': loading_rate
            })
            
            for level in transport_levels:
                level_df = date_df[date_df['运输等级'] == level]
                # 总体数据
                total_weight = level_df['计费重量(kg)'].sum()
                total_tickets = level_df['应到(件)'].sum()
                total_trips = level_df['计划需求ID'].nunique()
                total_full_weight = level_df['满载重量(kg)'].sum()
                low_load_trips = level_df[((level_df['线路里程（KM）'] > 800) & (level_df['重量装载率'] < 0.7)) | ((level_df['线路里程（KM）'] <= 800) & (level_df['重量装载率'] < 0.5))]['计划需求ID'].nunique()
                loading_rate = total_weight / total_full_weight if total_full_weight > 0 else 0

                # 存储总体数据
                results_data1.append({
                    '日期': date,
                    '片区': '分拨区',
                    '运输等级': level,
                    '发出网点': '总计',
                    '发运货量': total_weight,
                    '满载重量': total_full_weight,
                    '发运票数': total_tickets,
                    '发运车次': total_trips,
                    '低装载车次': low_load_trips,
                    '装载率': loading_rate
                })

            for region in fixed_regions:
                region_df = date_df[date_df['片区'] == region]

                # 总体数据
                total_weight = region_df['计费重量(kg)'].sum()
                total_tickets = region_df['应到(件)'].sum()
                total_trips = region_df['计划需求ID'].nunique()
                total_full_weight = region_df['满载重量(kg)'].sum()
                low_load_trips = region_df[((region_df['线路里程（KM）'] > 800) & (region_df['重量装载率'] < 0.7)) | ((region_df['线路里程（KM）'] <= 800) & (region_df['重量装载率'] < 0.5))]['计划需求ID'].nunique()
                loading_rate = total_weight / total_full_weight if total_full_weight > 0 else 0

                # 存储总体数据
                results_data1.append({
                    '日期': date,
                    '片区': region,
                    '运输等级': '总计',
                    '发出网点': '总计',
                    '发运货量': total_weight,
                    '满载重量': total_full_weight,
                    '发运票数': total_tickets,
                    '发运车次': total_trips,
                    '低装载车次': low_load_trips,
                    '装载率': loading_rate
                })

                for level in transport_levels:
                    level_df = region_df[region_df['运输等级'] == level]

                    # 发运货量
                    level_weight = level_df['计费重量(kg)'].sum()

                    # 发运票数
                    level_tickets = level_df['应到(件)'].sum()

                    # 发运车次
                    level_trips = level_df['计划需求ID'].nunique()

                    # 满载重量
                    level_full_weight = level_df['满载重量(kg)'].sum()

                    # 低装载车次
                    level_low_load_trips = level_df[((level_df['线路里程（KM）'] > 800) & (level_df['重量装载率'] < 0.7)) | ((level_df['线路里程（KM）'] <= 800) & (level_df['重量装载率'] < 0.5))]['计划需求ID'].nunique()

                    # 装载率
                    level_loading_rate = level_weight / level_full_weight if level_full_weight > 0 else 0

                    # 存储数据
                    results_data1.append({
                        '日期': date,
                        '片区': region,
                        '运输等级': level,
                        '发出网点': '总计',
                        '发运货量': level_weight,
                        '满载重量': level_full_weight,
                        '发运票数': level_tickets,
                        '发运车次': level_trips,
                        '低装载车次': level_low_load_trips,
                        '装载率': level_loading_rate
                    })

        # 创建 DataFrame
        results_df1 = pd.DataFrame(results_data1, columns=[
            '日期', '片区', '运输等级', '发出网点', '发运货量','满载重量', '发运票数', '发运车次', '低装载车次', '装载率'
        ])

        # 获取所有发出网点
        all_outlets = df['发出网点'].unique()

        fixed_outlet_order =['769XG','769WM','769WF','769WB','758CC001','757WA','757WE','763X','020WM','020RD','020W','020WE','020XB','755X','755WM','755WJ','755WE','755W','755WF','755RC','663WA','663WB','752WA','762VA','750VA','750W','760W','756W','759VA']

        # 按照发出网点分类计算统计数据
        results_data2 = []

        for date in all_dates:
            date_df = df[df['日期'].dt.date == date]

            # 总体数据
            total_weight = date_df['计费重量(kg)'].sum()
            total_tickets = date_df['应到(件)'].sum()
            total_trips = date_df['计划需求ID'].nunique()
            total_full_weight = date_df['满载重量(kg)'].sum()
            low_load_trips = date_df[((date_df['线路里程（KM）'] > 800) & (date_df['重量装载率'] < 0.7)) | ((date_df['线路里程（KM）'] <= 800) & (date_df['重量装载率'] < 0.5))]['计划需求ID'].nunique()
            loading_rate = total_weight / total_full_weight if total_full_weight > 0 else 0

            # 存储总体数据
            results_data2.append({
                '日期': date,
                '运输等级': '总计',
                '发出网点': '总计',
                '发运货量': total_weight,
                '满载重量': total_full_weight,
                '发运票数': total_tickets,
                '发运车次': total_trips,
                '低装载车次': low_load_trips,
                '装载率': loading_rate
            })

            for outlet in fixed_outlet_order:
                outlet_df = date_df[date_df['发出网点'] == outlet]

                # 总体数据
                outlet_weight = outlet_df['计费重量(kg)'].sum()
                outlet_tickets = outlet_df['应到(件)'].sum()
                outlet_trips = outlet_df['计划需求ID'].nunique()
                outlet_full_weight = outlet_df['满载重量(kg)'].sum()
                outlet_low_load_trips = outlet_df[((date_df['线路里程（KM）'] > 800) & (outlet_df['重量装载率'] < 0.7)) | ((outlet_df['线路里程（KM）'] <= 800) & (outlet_df['重量装载率'] < 0.5))]['计划需求ID'].nunique()
                outlet_loading_rate = outlet_weight / outlet_full_weight if outlet_full_weight > 0 else 0

                # 存储数据
                results_data2.append({
                    '日期': date,
                    '运输等级': '总计',
                    '发出网点': outlet,
                    '发运货量': outlet_weight,
                    '满载重量': outlet_full_weight,
                    '发运票数': outlet_tickets,
                    '发运车次': outlet_trips,
                    '低装载车次': outlet_low_load_trips,
                    '装载率': outlet_loading_rate
                })

                for level in transport_levels:
                    level_df = outlet_df[outlet_df['运输等级'] == level]

                    # 发运货量
                    level_weight = level_df['计费重量(kg)'].sum()

                    # 发运票数
                    level_tickets = level_df['应到(件)'].sum()

                    # 发运车次
                    level_trips = level_df['计划需求ID'].nunique()

                    # 满载重量
                    level_full_weight = level_df['满载重量(kg)'].sum()

                    # 低装载车次
                    level_low_load_trips = level_df[((level_df['线路里程（KM）'] > 800) & (level_df['计费重量(kg)'] / level_df['满载重量(kg)'] < 0.7)) | ((level_df['线路里程（KM）'] <= 800) & (level_df['计费重量(kg)'] / level_df['满载重量(kg)'] < 0.5))]['计划需求ID'].nunique()

                    # 装载率
                    level_loading_rate = level_weight / level_full_weight if level_full_weight > 0 else 0

                    # 存储数据
                    results_data2.append({
                        '日期': date,
                        '运输等级': level,
                        '发出网点': outlet,
                        '发运货量': level_weight,
                        '满载重量': level_full_weight,
                        '发运票数': level_tickets,
                        '发运车次': level_trips,
                        '低装载车次': level_low_load_trips,
                        '装载率': level_loading_rate
                    })

        # 创建 DataFrame
        results_df2 = pd.DataFrame(results_data2, columns=[
            '日期', '运输等级', '发出网点', '发运货量','满载重量', '发运票数', '发运车次', '低装载车次', '装载率'
        ])
        load_data_yesterday = (results_df1.loc[0:20,['片区','装载率']].reset_index(drop=True)).iloc[::3]
        load_data_today = (results_df1.loc[21:41,['片区','装载率']].reset_index(drop=True)).iloc[::3]
        load_rate1=(load_data_today['装载率']/load_data_yesterday['装载率']-1)
        load_data_today['环比'] = load_rate1

        load_data_area = []
        for index, load_area in enumerate(load_data_today['环比'][1:], start=1):
            if load_area < 0:
                load_data_area.append((load_data_today.iloc[index]['片区'], load_data_today.iloc[index]['环比']))

        load_data_area_df = pd.DataFrame(load_data_area, columns=['片区', '环比'])

        load_data_point_yesterday = (results_df2.loc[1:87,['发出网点','装载率']].reset_index(drop=True)).iloc[::3]
        load_data_point_today = (results_df2.loc[89:175,['发出网点','装载率']].reset_index(drop=True)).iloc[::3]
        load_rate2=(load_data_point_today['装载率']/load_data_point_yesterday['装载率']-1)
        load_data_point_today['环比'] = load_rate2

        # 定义网点别名映射
        网点别名映射 = {
            '769XG':'莞陆',
            '769WM':'常平',
            '769WF':'沙田',
            '769WB':'大朗',
            '758CC001':'唯品会',
            '757WA':'官窑',
            '757WE':'盐步',
            '763X':'清陆',
            '020WM':'化龙',
            '020RD':'穗航',
            '020W':'新塘',
            '020WE':'穗北',
            '020XB':'石龙',
            '755X':'深中',
            '755WM':'和记',
            '755WJ':'和记6号库',
            '755WE':'福永',
            '755W':'黄田',
            '755WF':'五和',
            '755RC':'深航二期',
            '663WA':'揭阳',
            '663WB':'揭阳A2库',
            '752WA':'惠州',
            '762VA':'河源',
            '750VA':'外海',
            '750W':'江门',
            '760W':'中山',
            '756W':'珠海',
            '759VA':'湛江'
        }

        # 筛选数据并替换网点名称
        load_data_point = []
        for index, load_point in enumerate(load_data_point_today['环比'][1:], start=1):
            if load_point < 0:
                发出网点 = load_data_point_today.iloc[index]['发出网点']
                环比 = load_data_point_today.iloc[index]['环比']
                if 发出网点 in 网点别名映射:
                    发出网点 = 网点别名映射[发出网点]
                load_data_point.append((发出网点, 环比))

        # 创建 DataFrame 并指定列名别名
        load_data_point_df = pd.DataFrame(load_data_point, columns=['网点', '环比'])



        from openpyxl import load_workbook

        # 加载现有的Excel工作簿
        workbook = load_workbook(filename='D:\\user\\01440394\\桌面\\规划部\\干线装载率\\干线达成预警test.xlsx')
        sheet = workbook['达成']  # 或者指定具体的工作表名称
        current_date = datetime.now().strftime('%m月%d日')
        direction = "上升" if load_data_today['环比'][0] > 0 else "下降"
        sheet.cell(row=4,column=3).value = f'①{current_date}截止{selected_time.hour, selected_time.minute}时分，分拨区小件重量装载率达成{results_df1.loc[21,'装载率']:.1%}，环比昨日同时段{direction}{load_data_today['环比'][0]:.1%}'
        load_data_area_df = load_data_area_df.sort_values(by='环比')
        load_data_point_df = load_data_point_df.sort_values(by='环比')
        formatted_data = load_data_area_df[['片区', '环比']].apply(lambda x: f"{x['片区']} ({x['环比']:.1%})", axis=1).tolist()
        formatted_point_data = load_data_point_df[['网点', '环比']].apply(lambda x: f"{x['网点']} ({x['环比']:.1%})", axis=1).tolist()
        if formatted_data:
            sheet.cell(row=28, column=3).value = f'②环比昨日同时段重量装载率向差片区：{", ".join(formatted_data)}\n ③环比昨日同时段重量装载率向差场地：{", ".join(formatted_point_data)}'
        else:
            sheet.cell(row=28, column=3).value = f'②环比昨日同时段重量装载率向差场地：{", ".join(formatted_point_data)}'
        

        # 选择"发运货量"列，并获取第0行到第10行的数据
        shipments = results_df1.loc[0:20, '发运货量']
        # 将数据写入E9到E19
        for index, value in enumerate(shipments):
            sheet.cell(row=6 + index + 1, column=5).value = value
            
        shipments = results_df1.loc[0:20,'满载重量']
        for index, value in enumerate(shipments):
            sheet.cell(row=6 + index + 1, column=6).value = value
            
            
        shipments = results_df1.loc[0:20,'发运车次']
        for index, value in enumerate(shipments):
            sheet.cell(row=6 + index + 1, column=7).value = value

            
        shipments = results_df1.loc[0:20,'低装载车次']
        for index, value in enumerate(shipments):
            sheet.cell(row=6 + index + 1, column=8).value = value
            
        shipments = results_df1.loc[0:20,'发运票数']
        for index, value in enumerate(shipments):
            sheet.cell(row=6 + index + 1, column=9).value = value
            
            
        shipments = results_df1.loc[0:20,'装载率']
        for index, value in enumerate(shipments):
            sheet.cell(row=6 + index + 1, column=10).value = value
            

        shipments = results_df1.loc[21:41,'发运货量']
        for index, value in enumerate(shipments):
            sheet.cell(row=6 + index + 1, column=11).value = value

        shipments = results_df1.loc[21:41,'满载重量']
        for index, value in enumerate(shipments):
            sheet.cell(row=6 + index + 1, column=12).value = value

        shipments = results_df1.loc[21:41,'发运车次']
        for index, value in enumerate(shipments):
            sheet.cell(row=6 + index + 1, column=13).value = value

        shipments = results_df1.loc[21:41,'低装载车次']
        for index, value in enumerate(shipments):
            sheet.cell(row=6 + index + 1, column=14).value = value

        shipments = results_df1.loc[21:41,'发运票数']
        for index, value in enumerate(shipments):
            sheet.cell(row=6 + index + 1, column=15).value = value

        shipments = results_df1.loc[21:41,'装载率']
        for index, value in enumerate(shipments):
            sheet.cell(row=6 + index + 1, column=16).value = value
            


        shipments = results_df2.loc[1:87,'发运货量']
        for index, value in enumerate(shipments):
            sheet.cell(row=30 + index + 1, column=5).value = value

        shipments = results_df2.loc[1:87,'满载重量']
        for index, value in enumerate(shipments):
            sheet.cell(row=30 + index + 1, column=6).value = value
            
        shipments = results_df2.loc[1:87,'发运车次']
        for index, value in enumerate(shipments):
            sheet.cell(row=30 + index + 1, column=7).value = value
            
        shipments = results_df2.loc[1:87,'低装载车次']
        for index, value in enumerate(shipments):
            sheet.cell(row=30 + index + 1, column=8).value = value

        shipments = results_df2.loc[1:87,'发运票数']
        for index, value in enumerate(shipments):
            sheet.cell(row=30 + index + 1, column=9).value = value

        shipments = results_df2.loc[1:87,'装载率']
        for index, value in enumerate(shipments):
            sheet.cell(row=30 + index + 1, column=10).value = value

        shipments = results_df2.loc[89:175,'发运货量']
        for index, value in enumerate(shipments):
            sheet.cell(row=30 + index + 1, column=11).value = value
            
        shipments = results_df2.loc[89:175,'满载重量']
        for index, value in enumerate(shipments):
            sheet.cell(row=30 + index + 1, column=12).value = value

        shipments = results_df2.loc[89:175,'发运车次']
        for index, value in enumerate(shipments):
            sheet.cell(row=30 + index + 1, column=13).value = value

        shipments = results_df2.loc[89:175,'低装载车次']
        for index, value in enumerate(shipments):
            sheet.cell(row=30 + index + 1, column=14).value = value

        shipments = results_df2.loc[89:175,'发运票数']
        for index, value in enumerate(shipments):
            sheet.cell(row=30 + index + 1, column=15).value = value

        shipments = results_df2.loc[89:175,'装载率']
        for index, value in enumerate(shipments):
            sheet.cell(row=30 + index + 1, column=16).value = value


    # 保存文件
    save_filename = st.sidebar.text_input("输入保存文件名", value="干线达成预警")
    if st.sidebar.button("保存文件"):
        save_path = os.path.join('D:\\user\\01440394\\桌面\\规划部\\干线装载率', f'{save_filename}.xlsx')
        workbook.save(filename=save_path)
        st.success(f"文件已保存为: {save_path}")
